import pytest
import json
import jsonpath
import requests
import openpyxl


class Common:

    def __init__(self, PathOfXl, SheetName):
        global workbook, sheet
        workbook = openpyxl.load_workbook("C:\\Users\\SridharRaju\\Desktop\\API_tasks\\newStudentDetail.xlsx")
        sheet = workbook['Sheet1']

    def fetch_row_count(self):
        rows = sheet.max_row
        return rows

    def fetch_col_count(self):
        cols = sheet.max_column
        return cols

    def fetch_key_names(self):
        c = sheet.max_column
        lst = []
        for i in range(1, c + 1):
            cell = sheet.cell(row=1, column=i)
            lst.insert(i-1, cell.value)
        return lst


    def update_request_with_data(self, rowNumber, jsonRequest, keyList):
        c = sheet.max_column
        for i in range(1, c + 1):
            cell = sheet.cell(row=rowNumber, column=i)
            jsonRequest[keyList[i-1]] = cell.value
        return jsonRequest